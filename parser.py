"""
Reads the uploaded .xlsm and splits the SHIPPING QUEUE orders into
red (OVERDUE) and yellow (TODAY) lists, keeping only the requested columns.

The SHIP STATUS column (A) is formula-driven in the workbook, so we read the
*cached* values (data_only=True). If the file was never opened/saved in Excel
those cached values can be missing — we detect that and raise a clear error.
"""
import datetime
import openpyxl

import config


class ParseError(Exception):
    """Raised for anything the user can fix (wrong sheet, no cached values...)."""


def _fmt(value):
    """Turn a cell value into a clean string for Google Sheets."""
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        # ISBNs / quantities come back as e.g. 9788184951837.0 -> "9788184951837"
        return str(int(value))
    return str(value).strip()


def parse_workbook(path):
    """
    Return a dict:
        {
          "headers":     [...6 column labels...],
          "red":         [[...], ...],   # OVERDUE rows
          "yellow":      [[...], ...],   # TODAY rows
          "red_count":   int,
          "yellow_count":int,
        }
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001 - surface a friendly message
        raise ParseError(f"Could not open the workbook: {e}")

    if config.SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ParseError(
            f"Sheet '{config.SHEET_NAME}' not found. "
            f"Sheets in this file: {', '.join(wb.sheetnames)}"
        )

    ws = wb[config.SHEET_NAME]

    headers = [label for label, _ in config.COLUMN_MAP]
    col_numbers = [num for _, num in config.COLUMN_MAP]

    red, yellow = [], []
    seen_statuses = set()

    # Row 1 is the header row in the source sheet -> start at row 2.
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        status = row[0] if len(row) > 0 else None          # col A = SHIP STATUS
        order_id = row[4] if len(row) > 4 else None         # col E = ORDER ID
        if status:
            seen_statuses.add(str(status).strip().upper())
        if order_id in (None, ""):
            continue
        status_norm = str(status).strip().upper() if status else ""
        if status_norm == config.RED_STATUS.upper():
            bucket = red
        elif status_norm == config.YELLOW_STATUS.upper():
            bucket = yellow
        else:
            continue
        out_row = [
            _fmt(row[num - 1]) if len(row) >= num else ""
            for num in col_numbers
        ]
        bucket.append(out_row)

    wb.close()

    if not red and not yellow:
        wanted = {config.RED_STATUS.upper(), config.YELLOW_STATUS.upper()}
        if not (seen_statuses & wanted):
            raise ParseError(
                "No orders had SHIP STATUS = "
                f"'{config.RED_STATUS}' or '{config.YELLOW_STATUS}'. "
                "The SHIP STATUS column is formula-driven — open the file in "
                "Excel and press Save once so the values are cached, then "
                "re-upload."
            )
        raise ParseError("No matching orders were found in the sheet.")

    return {
        "headers": headers,
        "red": red,
        "yellow": yellow,
        "red_count": len(red),
        "yellow_count": len(yellow),
    }
