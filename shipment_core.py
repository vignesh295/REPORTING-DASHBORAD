"""
Shipment processing core — pure functions, testable without web/Sheets/Drive.

Given a delivered AWB's shipment manifest (order-id, tracking, ship-date, …) and
the current unshipped-orders report, it:
  * derives each order's lane from its SKU suffix,
  * drops cancelled orders (any order not in the unshipped report),
  * renders the Amazon Shipping-Confirmation flat file for the rest.
"""
import csv
import os

# Amazon Shipping-Confirmation flat-file columns (sheet "ShippingConfirmation").
AMAZON_COLS = [
    "order-id", "order-item-id", "quantity", "ship-date",
    "carrier-code", "carrier-name", "tracking-number", "ship-method",
]

# Lane from the SKU suffix. Check the specific suffixes before the bare "au".
# (A plain-ISBN SKU — ends in a digit — is always IND -> UAE, since IND -> AUS
#  uses the "au" suffix; there is no plain-digit AUS case.)
_SUFFIX_LANES = [
    ("uaebz", "USA → UAE"),
    ("uaegd", "UK → UAE"),
    ("aubz",  "USA → AUS"),
    ("augd",  "UK → AUS"),
    ("au",    "IND → AUS"),
]


def lane_from_sku(sku):
    """Return the lane (e.g. 'USA -> UAE') for an Amazon SKU, or None if unknown."""
    s = str(sku or "").strip().lower()
    if not s:
        return None
    for suffix, lane in _SUFFIX_LANES:
        if s.endswith(suffix):
            return lane
    if s[-1].isdigit():
        return "IND → UAE"
    return None


def _clean(v):
    """Coerce a spreadsheet cell to a clean, JSON-safe string.
    xlsx cells arrive as native types — dates as datetime (not JSON-serialisable),
    whole numbers as float (1.0) — so normalise them here at the source."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))          # 1.0 -> "1"; avoids scientific notation too
    return str(v).strip()


def read_table(path):
    """Read a manifest / unshipped report (.txt|.csv tab/comma, or .xlsx/.xls).
    Returns (headers, rows) where rows is a list of dicts keyed by header."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            head = f.read(4096)
            f.seek(0)
            delim = "\t" if "\t" in head.splitlines()[0] else ","
            grid = [row for row in csv.reader(f, delimiter=delim)]
    if not grid:
        return [], []
    headers = [_clean(h) for h in grid[0]]
    rows = []
    for raw in grid[1:]:
        if not any(_clean(c) for c in raw):
            continue
        rows.append({headers[i]: _clean(raw[i]) for i in range(min(len(headers), len(raw)))})
    return headers, rows


def unshipped_order_ids(rows, id_field="order-id"):
    """The set of order-ids still present in the unshipped report (still valid)."""
    return {_clean(r.get(id_field)) for r in rows if _clean(r.get(id_field))}


def reconcile(manifest_rows, unshipped_ids, id_field="order-id"):
    """Split a delivered AWB's manifest into (keep, cancelled_order_ids).
    An order in the manifest but NOT in the unshipped report was cancelled."""
    keep, cancelled = [], set()
    for r in manifest_rows:
        oid = _clean(r.get(id_field))
        if not oid:
            continue
        if oid in unshipped_ids:
            keep.append(r)
        else:
            cancelled.add(oid)
    return keep, cancelled


def amazon_confirmation(rows):
    """Render rows as the Amazon Shipping-Confirmation flat file (tab-separated)."""
    out = ["\t".join(AMAZON_COLS)]
    for r in rows:
        out.append("\t".join(_clean(r.get(c)) for c in AMAZON_COLS))
    return "\n".join(out) + "\n"


def build_confirmation(manifest_rows, unshipped_ids, id_field="order-id"):
    """Full step: reconcile a manifest against the unshipped report and render
    the Amazon file. Returns (file_text, summary)."""
    keep, cancelled = reconcile(manifest_rows, unshipped_ids, id_field)
    text = amazon_confirmation(keep)
    summary = {
        "manifest_orders": len({_clean(r.get(id_field)) for r in manifest_rows if _clean(r.get(id_field))}),
        "confirmed_rows": len(keep),
        "confirmed_orders": len({_clean(r.get(id_field)) for r in keep}),
        "cancelled_orders": sorted(cancelled),
    }
    return text, summary
