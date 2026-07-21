"""
Red/Yellow splitter — turn a per-lane order report into RED (overdue) and
YELLOW (due-today) rows for the RED/YELLOW sheets.

Source: the report's "SHIPPING QUEUE" tab. The "SHIP STATUS" column classifies
each order — OVERDUE -> red, TODAY -> yellow, UPCOMING -> skip. The lane comes
from the file name. Column mapping to the output sheets:

  ORDER DATE <- DATE        ISBN  <- SKU         LAST SHIP DATE   <- LAST SHIP DATE
  ORDER ID   <- ORDER ID    TITLE <- TITLE NAME  ACTUAL SHIP DATE <- ACTUAL SHIP DATE
  QTY        <- QTY         REASON <- STATUS     LATE BY (red)    <- DAYS (+LEFT / -LATE)
"""
import os
import re

import openpyxl

import config
import sheets
from shipment_core import _clean

SOURCE_TAB = "SHIPPING QUEUE"
_ORIGIN = {"IND": "India", "INDIA": "India", "USA": "USA", "UK": "UK"}

# Ops SKU-code file names (e.g. "AMAZON ORDER REPORT AUBZ …"): the code encodes
# origin (BZ=USA, GD=UK) + destination (AU=AUS, UAE=UAE). UK-destination files are
# left out per ops (only USA / UAE / AUS destinations are handled).
_CODE_LANES = {
    "UAEBZ": "USA → UAE", "UAEGD": "UK → UAE",
    "AUBZ": "USA → AUS", "AUGD": "UK → AUS",
    # India-origin + UK→USA codes: add here once confirmed
    # e.g. "AUIN": "India → AUS", "UAEIN": "India → UAE", ...
}


def _match(lanes, target):
    t = target.replace(" ", "").upper()
    return next((l for l in lanes if l.replace(" ", "").upper() == t), None)


def _skip_lane(lane):
    """UK-destination lanes are intentionally not processed here."""
    return lane.split("→")[-1].strip().upper() == "UK"


def lane_from_filename(name, lanes):
    """Match a report file name to a configured lane — either 'ORIGIN TO DEST'
    or an ops SKU code (AUBZ, UAEGD, …). UK-destination lanes are skipped."""
    up = name.upper()
    m = re.search(r"([A-Z]+)\s+TO\s+([A-Z]+)", up)
    if m:
        lane = _match(lanes, f"{_ORIGIN.get(m.group(1), m.group(1).title())} → {m.group(2)}")
        if lane and not _skip_lane(lane):
            return lane
    for code in sorted(_CODE_LANES, key=len, reverse=True):  # longest first (AUBZ before AU)
        if re.search(r"\b" + code + r"\b", up):
            lane = _match(lanes, _CODE_LANES[code])
            if lane and not _skip_lane(lane):
                return lane
    return None


def _read_tab(path, tab):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        real = next((s for s in names if s.strip().upper() == tab.strip().upper()), None)
        if real is None:
            # Raise (rather than return empty) so a wrong-format file is skipped,
            # never used to clear a lane tab.
            raise ValueError(f"no '{tab}' tab (tabs: {names})")
        grid = [list(r) for r in wb[real].iter_rows(values_only=True)]
    finally:
        wb.close()
    if not grid:
        return [], []
    headers = [_clean(c) for c in grid[0]]
    rows = []
    for raw in grid[1:]:
        if not any(_clean(c) for c in raw):
            continue
        rows.append({headers[i]: _clean(raw[i]) for i in range(min(len(headers), len(raw)))})
    return headers, rows


def _get(row, *names):
    low = {str(k).strip().lower(): v for k, v in row.items()}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    return ""


def split_order_report(path, tab=SOURCE_TAB):
    """Return (red_rows, yellow_rows) as dicts keyed by the output column names.
    OVERDUE -> red, TODAY -> yellow (by the SHIP STATUS column); UPCOMING skipped."""
    _headers, rows = _read_tab(path, tab)
    red, yellow = [], []
    for r in rows:
        if _get(r, "ORDER ID").strip().upper() == "ORDER ID":  # repeated header row
            continue
        ship_status = _get(r, "SHIP STATUS").upper()
        out = {
            "ORDER DATE": _get(r, "DATE", "ORDER DATE"),
            "ORDER ID": _get(r, "ORDER ID"),
            "ISBN": _get(r, "SKU", "ISBN"),
            "TITLE": _get(r, "TITLE NAME", "TITLE"),
            "QTY": _get(r, "QTY"),
            "LAST SHIP DATE": _get(r, "LAST SHIP DATE"),
            "ACTUAL SHIP DATE": _get(r, "ACTUAL SHIP DATE"),
            "REASON": _get(r, "STATUS"),
        }
        if "OVERDUE" in ship_status:
            out["LATE BY"] = _get(r, "DAYS (+LEFT / -LATE)", "DAYS", "LATE BY")
            red.append(out)
        elif "TODAY" in ship_status:
            yellow.append(out)
    return red, yellow


# ---------------------------------------------------------------------------
# Writing the split into the RED / YELLOW sheets (one tab per lane)
# ---------------------------------------------------------------------------
def _cells(row, header):
    """Map a split-row dict to a cell list aligned to the tab's header order
    (header names matched case-/space-insensitively)."""
    rn = {str(k).strip().lower(): v for k, v in row.items()}
    return ["" if rn.get(h.strip().lower()) is None else str(rn.get(h.strip().lower(), ""))
            for h in header]


def _replace_tab(sh, title, rows):
    """Clear a lane tab's data and rewrite it (header preserved)."""
    ws = sh.worksheet(title)
    header = ws.row_values(1)
    if not header:
        return 0
    ws.clear()
    ws.update([header] + [_cells(r, header) for r in rows], "A1", raw=True)
    return len(rows)


def _accumulate_tab(sh, title, rows, id_key="ORDER ID"):
    """Append rows into a history tab, skipping Order IDs already present."""
    ws = sh.worksheet(title)
    existing = ws.get_all_values()
    if not existing:
        return 0
    header = existing[0]
    idi = next((i for i, h in enumerate(header) if h.strip().lower() == id_key.strip().lower()), None)
    seen = set()
    if idi is not None:
        for r in existing[1:]:
            if idi < len(r) and str(r[idi]).strip():
                seen.add(str(r[idi]).strip())
    new = []
    for r in rows:
        oid = str(r.get(id_key, "")).strip()
        if oid and oid in seen:
            continue
        if oid:
            seen.add(oid)
        new.append(_cells(r, header))
    if new:
        ws.append_rows(new, value_input_option="RAW")
    return len(new)


def write_split(lane, red_rows, yellow_rows, red_id=None, yellow_id=None,
                history_suffix=" — Previous"):
    """Write a lane's split: refresh the current RED + YELLOW lane tabs and
    accumulate the RED rows into the lane's history tab (dedup by Order ID)."""
    shr = sheets.client().open_by_key(red_id or config.RED_SPREADSHEET_ID)
    shy = sheets.client().open_by_key(yellow_id or config.YELLOW_SPREADSHEET_ID)
    res = {"lane": lane,
           "yellow": _replace_tab(shy, lane, yellow_rows),
           "red": _replace_tab(shr, lane, red_rows)}
    try:
        res["history_added"] = _accumulate_tab(shr, lane + history_suffix, red_rows)
    except Exception:  # history tab missing -> skip, don't fail the write
        res["history_added"] = 0
    return res


def process_order_report(path, lanes=None):
    """Detect the lane from the file name, split it, and write to the sheets.
    Returns {'error': ...} (no write) if the lane can't be detected or the file
    isn't in the expected SHIPPING QUEUE format."""
    lanes = lanes or config.LANES
    lane = lane_from_filename(os.path.basename(path), lanes)
    if not lane:
        return {"error": f"no lane detected in {os.path.basename(path)!r}"}
    try:
        red, yellow = split_order_report(path)
    except Exception as e:  # noqa: BLE001  (missing tab / unreadable -> skip, don't clear)
        return {"error": f"{lane}: {e}", "lane": lane}
    return write_split(lane, red, yellow)
